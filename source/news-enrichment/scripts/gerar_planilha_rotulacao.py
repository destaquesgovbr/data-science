"""
Gerador de Planilha para Rotulação Manual

Cria uma planilha Excel formatada para rotulação manual de notícias.
Útil para criar dados de treino para BERT ou validar classificações.

Uso:
    poetry run python source/news-enrichment/scripts/gerar_planilha_rotulacao.py --noticias 100
"""

import argparse
import random
from pathlib import Path
from datetime import datetime
import polars as pl
import sys

# Configuração de paths
PROJECT_ROOT = Path(__file__).parent.parent.parent.parent
sys.path.insert(0, str(PROJECT_ROOT / "source" / "news-enrichment"))

DATA_DIR = PROJECT_ROOT / "data"

# Imports
import yaml


def load_taxonomy():
    """Carrega taxonomia completa."""
    arvore_path = PROJECT_ROOT / "arvore.yaml"

    with open(arvore_path, 'r', encoding='utf-8') as f:
        taxonomy = yaml.safe_load(f)

    return taxonomy


def extract_all_categories(taxonomy):
    """Extrai todas as categorias da taxonomia em formato flat."""
    categories = []

    def recurse(node, nivel_1='', nivel_2=''):
        if isinstance(node, dict):
            for key, value in node.items():
                if key == '_categorias':
                    # Categorias finais (folhas)
                    for cat in value:
                        categories.append({
                            'nivel_1': nivel_1,
                            'nivel_2': nivel_2,
                            'categoria': cat,
                            'path': f"{nivel_1} > {nivel_2} > {cat}" if nivel_2 else f"{nivel_1} > {cat}"
                        })
                elif nivel_1 == '':
                    # Nível 1
                    recurse(value, nivel_1=key)
                elif nivel_2 == '':
                    # Nível 2
                    recurse(value, nivel_1=nivel_1, nivel_2=key)
                else:
                    # Níveis mais profundos
                    recurse(value, nivel_1=nivel_1, nivel_2=nivel_2)
        elif isinstance(node, list):
            # Lista de categorias
            for cat in node:
                categories.append({
                    'nivel_1': nivel_1,
                    'nivel_2': nivel_2 if nivel_2 else '',
                    'categoria': cat,
                    'path': f"{nivel_1} > {nivel_2} > {cat}" if nivel_2 else f"{nivel_1} > {cat}"
                })

    recurse(taxonomy)
    return categories


def load_random_news(n: int = 100, seed: int = 42):
    """Carrega n notícias aleatórias do dataset."""
    print(f"\nCarregando {n} notícias aleatórias do dataset...")

    dataset_path = DATA_DIR / "govbrnews_full.parquet"

    if not dataset_path.exists():
        alternatives = [
            DATA_DIR / "sample_enriched.parquet",
            DATA_DIR / "sample_enriched_otimizado_500.parquet",
        ]

        for alt in alternatives:
            if alt.exists():
                dataset_path = alt
                print(f"Usando dataset alternativo: {alt.name}")
                break
        else:
            raise FileNotFoundError("Nenhum dataset de notícias encontrado")

    df = pl.read_parquet(dataset_path)

    # Filtrar válidas
    df = df.filter(
        (pl.col('title').is_not_null()) &
        (pl.col('content').is_not_null()) &
        (pl.col('content').str.len_chars() > 100)
    )

    total = len(df)
    print(f"✓ {total:,} notícias válidas no dataset")

    if total < n:
        print(f"⚠️  Dataset tem apenas {total} notícias (pedido: {n})")
        n = total

    # Selecionar aleatoriamente
    random.seed(seed)
    random_indices = random.sample(range(total), n)
    df = df[random_indices]

    print(f"✓ {n} notícias selecionadas aleatoriamente\n")

    return df


def create_labeling_sheet(df: pl.DataFrame, categories: list):
    """Cria planilha de rotulação."""
    print("Criando planilha de rotulação...")

    # Preparar dados
    rows = []

    for i, row in enumerate(df.iter_rows(named=True), 1):
        rows.append({
            'id': i,
            'unique_id': row.get('unique_id', f'news_{i}'),
            'titulo': row['title'],
            'conteudo': row['content'],
            'conteudo_preview': row['content'][:300].replace('\n', ' ') + '...',
            'nivel_1_rotulado': '',
            'nivel_2_rotulado': '',
            'categoria_rotulada': '',
            'confianca': '',  # Alta, Média, Baixa
            'observacoes': '',
            'rotulador': '',  # Nome de quem rotulou
            'data_rotulacao': '',
            'tempo_gasto_min': '',  # Tempo em minutos
        })

    df_labeling = pl.DataFrame(rows)

    # Criar DataFrame de categorias (para aba de referência)
    df_categories = pl.DataFrame(categories)

    print(f"✓ Planilha criada com {len(df_labeling)} notícias\n")

    return df_labeling, df_categories


def create_instructions_sheet():
    """Cria aba de instruções."""
    instructions = {
        'secao': [
            'OBJETIVO',
            'INSTRUÇÕES GERAIS',
            'PASSO A PASSO',
            'PASSO 1',
            'PASSO 2',
            'PASSO 3',
            'PASSO 4',
            'PASSO 5',
            'PASSO 6',
            'CAMPO CONFIANÇA',
            'DICAS',
            'DICA 1',
            'DICA 2',
            'DICA 3',
            'DICA 4',
            'ESTATÍSTICAS',
            'TEMPO ESTIMADO',
            'CONTATO',
        ],
        'conteudo': [
            'Rotular manualmente notícias para criar dataset de treino para modelos de ML',
            'Leia o título e conteúdo da notícia e escolha a categoria mais específica da taxonomia',
            'Para cada notícia:',
            'Ler título e conteúdo completo',
            'Consultar taxonomia na aba "Categorias" para encontrar categoria correta',
            'Preencher Nível 1, Nível 2 e Categoria mais específica',
            'Avaliar sua confiança: Alta (certeza), Média (razoável), Baixa (dúvida)',
            'Adicionar observações se necessário (ex: notícia ambígua, categoria não existe)',
            'Preencher seu nome e data (para controle)',
            'Alta = Certeza absoluta | Média = Razoavelmente confiante | Baixa = Incerto/ambíguo',
            'Sugestões para rotulação eficiente:',
            'Leia primeiro o título - geralmente indica o tema principal',
            'Se em dúvida, leia os primeiros 2-3 parágrafos do conteúdo',
            'Use Ctrl+F na aba Categorias para buscar palavras-chave',
            'Anote casos difíceis para discussão posterior com a equipe',
            'Para criar dataset de treino BERT:',
            'Mínimo: 20-50 exemplos por categoria | Ideal: 100+ exemplos por categoria | Tempo: ~2-5min por notícia | Total: 100 notícias = 3-8 horas',
            'Dúvidas ou problemas: contatar Luis Felipe de Moraes',
        ]
    }

    return pl.DataFrame(instructions)


def create_stats_sheet():
    """Cria aba de estatísticas (para preencher durante rotulação)."""
    stats = {
        'metrica': [
            'Total de notícias',
            'Notícias rotuladas',
            'Pendentes',
            '% Completo',
            '',
            'Confiança Alta',
            'Confiança Média',
            'Confiança Baixa',
            '',
            'Tempo total (min)',
            'Tempo médio por notícia (min)',
            '',
            'Rotuladores',
        ],
        'valor': [
            '=COUNTA(Rotulacao!A:A)-1',  # Total rows - header
            '=COUNTIF(Rotulacao!F:F,"<>")',  # Células preenchidas
            '=COUNTIF(Rotulacao!F:F,"")',  # Células vazias
            '=B2/B1',
            '',
            '=COUNTIF(Rotulacao!I:I,"Alta")',
            '=COUNTIF(Rotulacao!I:I,"Média")',
            '=COUNTIF(Rotulacao!I:I,"Baixa")',
            '',
            '=SUM(Rotulacao!M:M)',
            '=AVERAGE(Rotulacao!M:M)',
            '',
            '=COUNTA(UNIQUE(Rotulacao!K:K))-1',
        ],
        'observacao': [
            'Número total de notícias na planilha',
            'Notícias com categoria preenchida',
            'Notícias ainda sem rotulação',
            'Progresso da rotulação',
            '',
            'Rotulações com certeza alta',
            'Rotulações razoáveis',
            'Rotulações com dúvida',
            '',
            'Tempo total gasto (em minutos)',
            'Média de tempo por notícia',
            '',
            'Número de pessoas que rotularam',
        ]
    }

    return pl.DataFrame(stats)


def save_workbook(df_labeling, df_categories, df_instructions, df_stats, output_path):
    """Salva workbook com múltiplas abas."""
    print(f"Salvando planilha: {output_path.name}...")

    # Polars não suporta múltiplas abas diretamente, usar openpyxl
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment
    from openpyxl.utils.dataframe import dataframe_to_rows

    wb = Workbook()

    # Remover aba padrão
    wb.remove(wb.active)

    # 1. Aba de Instruções (primeira aba)
    ws_instructions = wb.create_sheet("Instruções", 0)
    ws_instructions.sheet_properties.tabColor = "00FF00"  # Verde

    for r_idx, row in enumerate(dataframe_to_rows(df_instructions.to_pandas(), index=False, header=True), 1):
        for c_idx, value in enumerate(row, 1):
            cell = ws_instructions.cell(row=r_idx, column=c_idx, value=value)

            # Formatar cabeçalho
            if r_idx == 1:
                cell.font = Font(bold=True, size=12)
                cell.fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
                cell.font = Font(bold=True, color="FFFFFF")
            # Formatar seções (coluna A)
            elif c_idx == 1:
                cell.font = Font(bold=True)
                cell.fill = PatternFill(start_color="E7E6E6", end_color="E7E6E6", fill_type="solid")

            cell.alignment = Alignment(wrap_text=True, vertical="top")

    ws_instructions.column_dimensions['A'].width = 25
    ws_instructions.column_dimensions['B'].width = 80

    # 2. Aba de Rotulação (segunda aba)
    ws_labeling = wb.create_sheet("Rotulacao", 1)
    ws_labeling.sheet_properties.tabColor = "FFC000"  # Laranja

    for r_idx, row in enumerate(dataframe_to_rows(df_labeling.to_pandas(), index=False, header=True), 1):
        for c_idx, value in enumerate(row, 1):
            cell = ws_labeling.cell(row=r_idx, column=c_idx, value=value)

            # Formatar cabeçalho
            if r_idx == 1:
                cell.font = Font(bold=True, size=11)
                cell.fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
                cell.font = Font(bold=True, color="FFFFFF")
                cell.alignment = Alignment(horizontal="center")
            # Formatar células de preenchimento (colunas F-M)
            elif c_idx >= 6:
                cell.fill = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")

            cell.alignment = Alignment(wrap_text=True, vertical="top")

    # Ajustar larguras
    ws_labeling.column_dimensions['A'].width = 5   # ID
    ws_labeling.column_dimensions['B'].width = 15  # unique_id
    ws_labeling.column_dimensions['C'].width = 40  # titulo
    ws_labeling.column_dimensions['D'].width = 60  # conteudo (ocultar)
    ws_labeling.column_dimensions['E'].width = 50  # preview
    ws_labeling.column_dimensions['F'].width = 20  # nivel_1
    ws_labeling.column_dimensions['G'].width = 25  # nivel_2
    ws_labeling.column_dimensions['H'].width = 30  # categoria
    ws_labeling.column_dimensions['I'].width = 12  # confianca
    ws_labeling.column_dimensions['J'].width = 40  # observacoes
    ws_labeling.column_dimensions['K'].width = 20  # rotulador
    ws_labeling.column_dimensions['L'].width = 15  # data
    ws_labeling.column_dimensions['M'].width = 10  # tempo

    # Ocultar coluna de conteúdo completo (muito longa)
    ws_labeling.column_dimensions['D'].hidden = True

    # Congelar primeira linha e primeiras 3 colunas
    ws_labeling.freeze_panes = 'D2'

    # 3. Aba de Categorias (terceira aba)
    ws_categories = wb.create_sheet("Categorias", 2)
    ws_categories.sheet_properties.tabColor = "00B0F0"  # Azul

    for r_idx, row in enumerate(dataframe_to_rows(df_categories.to_pandas(), index=False, header=True), 1):
        for c_idx, value in enumerate(row, 1):
            cell = ws_categories.cell(row=r_idx, column=c_idx, value=value)

            if r_idx == 1:
                cell.font = Font(bold=True, size=11)
                cell.fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
                cell.font = Font(bold=True, color="FFFFFF")

            cell.alignment = Alignment(wrap_text=True, vertical="top")

    ws_categories.column_dimensions['A'].width = 25  # nivel_1
    ws_categories.column_dimensions['B'].width = 30  # nivel_2
    ws_categories.column_dimensions['C'].width = 35  # categoria
    ws_categories.column_dimensions['D'].width = 60  # path completo

    # 4. Aba de Estatísticas (quarta aba)
    ws_stats = wb.create_sheet("Estatisticas", 3)
    ws_stats.sheet_properties.tabColor = "92D050"  # Verde claro

    for r_idx, row in enumerate(dataframe_to_rows(df_stats.to_pandas(), index=False, header=True), 1):
        for c_idx, value in enumerate(row, 1):
            cell = ws_stats.cell(row=r_idx, column=c_idx, value=value)

            if r_idx == 1:
                cell.font = Font(bold=True, size=11)
                cell.fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
                cell.font = Font(bold=True, color="FFFFFF")
            elif c_idx == 1:
                cell.font = Font(bold=True)

            cell.alignment = Alignment(wrap_text=True, vertical="top")

    ws_stats.column_dimensions['A'].width = 30
    ws_stats.column_dimensions['B'].width = 20
    ws_stats.column_dimensions['C'].width = 50

    # Salvar
    wb.save(output_path)
    print(f"✓ Planilha salva: {output_path}\n")


def main():
    parser = argparse.ArgumentParser(description='Gera planilha para rotulação manual de notícias')
    parser.add_argument('--noticias', type=int, default=100, help='Número de notícias (padrão: 100)')
    parser.add_argument('--seed', type=int, default=42, help='Seed para reprodutibilidade (padrão: 42)')

    args = parser.parse_args()

    print("\n" + "="*80)
    print("GERADOR DE PLANILHA PARA ROTULAÇÃO MANUAL")
    print("="*80)
    print()
    print(f"Configuração:")
    print(f"  Notícias: {args.noticias}")
    print(f"  Seed: {args.seed}")
    print()

    # 1. Carregar taxonomia
    print("Carregando taxonomia...")
    taxonomy = load_taxonomy()
    categories = extract_all_categories(taxonomy)
    print(f"✓ {len(categories)} categorias na taxonomia\n")

    # 2. Carregar notícias
    df = load_random_news(n=args.noticias, seed=args.seed)

    # 3. Criar planilha de rotulação
    df_labeling, df_categories = create_labeling_sheet(df, categories)

    # 4. Criar instruções
    df_instructions = create_instructions_sheet()

    # 5. Criar estatísticas
    df_stats = create_stats_sheet()

    # 6. Salvar
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    output_path = DATA_DIR / f"rotulacao_manual_{args.noticias}_noticias_{timestamp}.xlsx"

    save_workbook(df_labeling, df_categories, df_instructions, df_stats, output_path)

    # Resumo
    print("="*80)
    print("✓ PLANILHA GERADA COM SUCESSO!")
    print("="*80)
    print()
    print(f"Arquivo: {output_path.name}")
    print()
    print("Abas criadas:")
    print("  1. Instruções - Leia primeiro!")
    print("  2. Rotulacao - Preencher aqui")
    print("  3. Categorias - Consultar taxonomia")
    print("  4. Estatisticas - Acompanhar progresso")
    print()
    print("Próximos passos:")
    print(f"  1. Abrir: {output_path}")
    print("  2. Ler aba 'Instruções'")
    print("  3. Rotular notícias na aba 'Rotulacao'")
    print("  4. Acompanhar progresso na aba 'Estatisticas'")
    print()
    print(f"Tempo estimado: {args.noticias * 3}-{args.noticias * 5} minutos")
    print(f"                (~{args.noticias * 3 / 60:.1f}-{args.noticias * 5 / 60:.1f} horas)")
    print()


if __name__ == "__main__":
    main()
