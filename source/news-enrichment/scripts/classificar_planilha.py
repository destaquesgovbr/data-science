"""
Classificador Automático de Planilha de Rotulação

Classifica notícias da planilha de rotulação usando Claude e preenche
automaticamente todos os campos.

Útil para:
1. Criar dados rotulados rapidamente
2. Treinar BERT como demonstração
3. Mostrar processo completo ao gestor

Uso:
    # Classificar planilha mais recente
    poetry run python source/news-enrichment/scripts/classificar_planilha.py

    # Especificar arquivo
    poetry run python source/news-enrichment/scripts/classificar_planilha.py --arquivo data/rotulacao_manual_20_noticias_20260224_120000.xlsx
"""

import argparse
from pathlib import Path
from datetime import datetime
import polars as pl
import sys
import time

# Configuração de paths
PROJECT_ROOT = Path(__file__).parent.parent.parent.parent
sys.path.insert(0, str(PROJECT_ROOT / "source" / "news-enrichment"))

DATA_DIR = PROJECT_ROOT / "data"

# Imports
from news_enrichment import NewsClassifier


def find_latest_labeling_sheet():
    """Encontra planilha de rotulação mais recente."""
    files = list(DATA_DIR.glob("rotulacao_manual_*.xlsx"))

    if not files:
        return None

    # Ordenar por data de modificação
    files.sort(key=lambda x: x.stat().st_mtime, reverse=True)
    return files[0]


def classify_and_fill_sheet(file_path: Path):
    """Classifica notícias e preenche planilha."""
    print("\n" + "="*80)
    print("CLASSIFICAÇÃO AUTOMÁTICA DA PLANILHA")
    print("="*80)
    print()
    print(f"Arquivo: {file_path.name}\n")

    # 1. Carregar planilha
    print("Carregando planilha...")
    df = pl.read_excel(file_path, sheet_name="Rotulacao")
    print(f"✓ {len(df)} notícias na planilha\n")

    # Verificar quantas já estão rotuladas
    rotuladas = df.filter(pl.col('categoria_rotulada') != '').height
    pendentes = len(df) - rotuladas

    if rotuladas > 0:
        print(f"⚠️  {rotuladas} notícias já rotuladas")
        print(f"   Classificando apenas {pendentes} pendentes\n")

    if pendentes == 0:
        print("✓ Todas as notícias já estão rotuladas!")
        return df

    # 2. Inicializar classificador
    print("Inicializando classificador Claude...")
    start_init = time.time()
    classifier = NewsClassifier(verbose=False)
    init_time = time.time() - start_init
    print(f"✓ Inicializado em {init_time:.2f}s\n")

    # 3. Classificar cada notícia pendente
    print(f"Classificando {pendentes} notícias...")
    print("-" * 80)

    results = []
    start_classify = time.time()

    for i, row in enumerate(df.iter_rows(named=True), 1):
        # Pular se já rotulada
        if row['categoria_rotulada'] and row['categoria_rotulada'] != '':
            results.append(row)
            continue

        print(f"[{i:2d}/{len(df)}] {row['titulo'][:60]}...")

        # Criar objeto notícia
        news = {
            'unique_id': row['unique_id'],
            'title': row['titulo'],
            'content': row['conteudo']
        }

        # Classificar
        try:
            start_single = time.time()
            result = classifier.classify_single(news, return_format="dict")
            elapsed = time.time() - start_single

            # Preencher campos
            row_updated = {
                **row,
                'nivel_1_rotulado': result.get('theme_1_level_1_label', ''),
                'nivel_2_rotulado': result.get('theme_1_level_2_label', ''),
                'categoria_rotulada': result.get('most_specific_theme_label', ''),
                'confianca': 'Alta',
                'observacoes': f'Rotulado automaticamente por Claude ({elapsed:.1f}s)',
                'rotulador': 'Claude (automático)',
                'data_rotulacao': datetime.now().strftime('%Y-%m-%d'),
                'tempo_gasto_min': f'{elapsed/60:.2f}'
            }

            results.append(row_updated)
            print(f"          ✓ {result.get('most_specific_theme_label', 'N/A')}")

        except Exception as e:
            print(f"          ✗ Erro: {e}")
            row_error = {
                **row,
                'observacoes': f'ERRO na classificação: {str(e)[:100]}'
            }
            results.append(row_error)

    classify_time = time.time() - start_classify

    print("-" * 80)
    print(f"✓ Classificação concluída em {classify_time:.1f}s")
    print(f"  Média: {classify_time/pendentes:.2f}s por notícia\n")

    # 4. Criar DataFrame atualizado
    df_updated = pl.DataFrame(results)

    return df_updated


def save_filled_sheet(df: pl.DataFrame, original_path: Path):
    """Salva planilha preenchida."""
    print("Salvando planilha preenchida...")

    # Nome do arquivo de saída
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    output_path = original_path.parent / f"{original_path.stem}_PREENCHIDO_{timestamp}.xlsx"

    # Converter para pandas (openpyxl funciona melhor com pandas)
    df_pandas = df.to_pandas()

    # Salvar com formatação básica
    from openpyxl import load_workbook
    from openpyxl.styles import Font, PatternFill

    # Salvar DataFrame
    df_pandas.to_excel(output_path, sheet_name="Rotulacao", index=False)

    # Carregar e formatar
    wb = load_workbook(output_path)
    ws = wb["Rotulacao"]

    # Formatar cabeçalho
    for cell in ws[1]:
        cell.font = Font(bold=True, size=11)
        cell.fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        cell.font = Font(bold=True, color="FFFFFF")

    # Destacar células preenchidas automaticamente (verde claro)
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=6, max_col=13):
        for cell in row:
            if cell.value and cell.value != '':
                cell.fill = PatternFill(start_color="D9EAD3", end_color="D9EAD3", fill_type="solid")

    # Ajustar larguras
    ws.column_dimensions['A'].width = 5
    ws.column_dimensions['B'].width = 15
    ws.column_dimensions['C'].width = 40
    ws.column_dimensions['D'].width = 60
    ws.column_dimensions['E'].width = 50
    ws.column_dimensions['F'].width = 20
    ws.column_dimensions['G'].width = 25
    ws.column_dimensions['H'].width = 30
    ws.column_dimensions['I'].width = 12
    ws.column_dimensions['J'].width = 40
    ws.column_dimensions['K'].width = 20
    ws.column_dimensions['L'].width = 15
    ws.column_dimensions['M'].width = 10

    # Ocultar coluna de conteúdo completo
    ws.column_dimensions['D'].hidden = True

    # Congelar painéis
    ws.freeze_panes = 'D2'

    wb.save(output_path)

    print(f"✓ Planilha salva: {output_path.name}\n")

    return output_path


def export_to_training_format(df: pl.DataFrame, output_path: Path):
    """Exporta dados no formato para treino BERT."""
    print("Exportando para formato de treino BERT...")

    # Filtrar apenas rotuladas
    df_train = df.filter(
        (pl.col('categoria_rotulada') != '') &
        (pl.col('categoria_rotulada').is_not_null())
    )

    if len(df_train) == 0:
        print("⚠️  Nenhuma notícia rotulada para exportar")
        return None

    # Preparar formato
    df_export = df_train.select([
        pl.col('titulo').alias('title'),
        pl.col('conteudo').alias('content'),
        pl.col('nivel_1_rotulado').alias('level_1'),
        pl.col('nivel_2_rotulado').alias('level_2'),
        pl.col('categoria_rotulada').alias('category'),
    ])

    # Salvar como parquet
    train_path = output_path.parent / f"dataset_treino_bert_{datetime.now().strftime('%Y%m%d_%H%M%S')}.parquet"
    df_export.write_parquet(train_path)

    print(f"✓ Dataset de treino salvo: {train_path.name}")
    print(f"  {len(df_export)} notícias prontas para BERT\n")

    # Estatísticas
    print("Distribuição de categorias:")
    cat_counts = df_export.group_by('category').len().sort('len', descending=True).head(10)
    for row in cat_counts.iter_rows(named=True):
        print(f"  {row['category']:40s}: {row['len']:2d}")

    return train_path


def main():
    parser = argparse.ArgumentParser(description='Classifica notícias da planilha de rotulação')
    parser.add_argument('--arquivo', type=str, help='Caminho para planilha Excel')
    parser.add_argument('--exportar', action='store_true', help='Exportar para formato de treino BERT')

    args = parser.parse_args()

    # Encontrar arquivo
    if args.arquivo:
        file_path = Path(args.arquivo)
        if not file_path.exists():
            print(f"✗ Arquivo não encontrado: {file_path}")
            return
    else:
        file_path = find_latest_labeling_sheet()
        if not file_path:
            print("✗ Nenhuma planilha de rotulação encontrada em data/")
            print("\nGere uma planilha primeiro:")
            print("  poetry run python source/news-enrichment/scripts/gerar_planilha_rotulacao.py")
            return

        print(f"Usando planilha mais recente: {file_path.name}\n")

    # Classificar
    df_filled = classify_and_fill_sheet(file_path)

    # Salvar
    output_path = save_filled_sheet(df_filled, file_path)

    # Exportar para treino (opcional)
    if args.exportar:
        export_to_training_format(df_filled, output_path)

    # Resumo
    print("="*80)
    print("✓ CLASSIFICAÇÃO CONCLUÍDA!")
    print("="*80)
    print()
    print("Arquivos gerados:")
    print(f"  1. {output_path.name}")
    print(f"     → Planilha preenchida com classificações")

    if args.exportar:
        print(f"  2. dataset_treino_bert_*.parquet")
        print(f"     → Formato pronto para treinar BERT")

    print()
    print("Próximos passos:")
    print(f"  1. Revisar: {output_path}")
    print("  2. (Opcional) Corrigir classificações manualmente")

    if args.exportar:
        print("  3. Treinar BERT:")
        print("     poetry run python source/news-enrichment/train_bert_classifier.py")
    else:
        print("  3. Exportar para treino:")
        print("     poetry run python source/news-enrichment/scripts/classificar_planilha.py --exportar")

    print()


if __name__ == "__main__":
    main()
